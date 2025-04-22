from math import atan, degrees
from tkinter import filedialog as dlg
import cv2
import numpy as np
import pandas as pd
import os
import math
import multiprocessing
from openpyxl import Workbook  # pip install openpyxl
from openpyxl import load_workbook
from skimage.measure import label, regionprops_table
from skimage.transform import resize
from skimage.measure import find_contours
from skimage.color import rgb2hsv
from skimage.morphology import disk
from skimage.filters import median
from scipy.optimize import curve_fit  # Lib p/ fazer ajuste polinomial
import matplotlib.pyplot as plt
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor
import time
from functools import wraps

def timing(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        start = time.perf_counter()
        result = func(*args, **kwargs)
        end = time.perf_counter()
        print(f"[TIMER] {func.__name__} levou {end-start:.3f}s")
        return result
    return wrapper

# --------------------------------------------------------------------
# 1) CPU‑only: roda em subprocessos, sem NENHUM I/O
# --------------------------------------------------------------------
def cpu_process_egg(job):
    """
    Recebe um job com (egg_num, rec_bytes, shape, l0,l1,c0,c1,pixFactor),
    executa SOMENTE o processamento numérico e devolve tudo em memória.
    """
    egg_num, rec_bytes, shape, l0, l1, c0, c1, pixFactor = job
    rec = np.frombuffer(rec_bytes, dtype=np.uint8).reshape(shape)
    out = process(rec, 1, pixFactor, None, None, None)
    # agora out inclui também pAi, pAf, pBi, pBf
    original, A, B, C, D, chosen_vol, chosen_area, x_data, y_data, pAi, pAf, pBi, pBf = out
    return egg_num, l0, l1, c0, c1, (original, A, B, C, D, chosen_vol, chosen_area, x_data, y_data, pAi, pAf, pBi, pBf)

def check_and_create_directory_if_not_exist(path_directory):
    if not os.path.exists(path_directory):
        os.makedirs(path_directory)
        print(f"The path {path_directory} is created!")

# Esta função ajusta as dimensões de uma imagem para garantir que ela
# não exceda 1800 pixels em largura ou altura, mantendo a proporção.
# im > imagem de entrada
def adjustImageDimension(im):
    tLines, tColumns, c = im.shape
    if tLines > tColumns:
        rFactor = 1800 / tLines
        tLines = 1800
        tColumns = int(tColumns * rFactor)
    else:
        rFactor = 1800 / tColumns
        tColumns = 1800
        tLines = int(tLines * rFactor)
    return (tLines, tColumns, rFactor)

# Esta função desenha linhas na imagem com base em coordenadas fornecidas para linhas
# de calibração, temporárias e de elementos.
# variaveis: image -> imagem na qual as linhas serão desenhadas
# calibrationLine -> coordenadas da linha de calibração
# temLines -> lista de coordenadas para linhas temp
# elementLines -> lista de coordenadas para linhas de elementos
# totalColumns -> numero total de colunas a srem desenhadas
# totalLines -> numero total de linhas a serem desenhadas
def drawLines(image, calibrationLine, tempLines, elementLines, totalColumns, totalLines):
    if len(calibrationLine) > 0:
        xi, yi, c = calibrationLine[0]
        xf, yf, c = calibrationLine[1]
        cv2.line(image, (xi, yi), (xf, yf), (0, 0, 255), 1)
    else:
        if len(tempLines) > 0:
            (posTemp, status) = tempLines[0]
            if status == 1:  # indica uma coluna
                cv2.line(image, (0, posTemp), (totalColumns - 1, posTemp), (0, 255, 0), 1)
            elif status == 2:
                cv2.line(image, (posTemp, 0), (posTemp, totalLines - 1), (255, 0, 0), 1)

        for i in elementLines:
            (posTemp, status) = i
            if status == 1:  # indica uma coluna
                cv2.line(image, (0, posTemp), (totalColumns - 1, posTemp), (0, 255, 0), 1)
            elif status == 2:
                cv2.line(image, (posTemp, 0), (posTemp, totalLines - 1), (255, 0, 0), 1)

    return (image)

# Essa função extrai sub-imagens de uma imagem maior com base
# em linhas temporárias e de elementos.
# variaveis: image -> imagem da qual as subimagens serão extraidas
# temLines -> lista de coordenadas para linhas temp
# elementLines -> lista de coordenadas para linhas de elementos
@timing
def subImagens(img, tColumns, tLines, factor, pixFactor, dFactor, nomeArquivo):
    t0 = time.perf_counter()
    """
    1) monta a lista de jobs (imagem, coordenadas, pixFactor)
    2) paraleliza SÓ cpu_process_egg (sem I/O interno)
    3) fora do pool, faz TODO o I/O:
       - grava PNGs
       - escreve Relatorio.dad
       - gera gráficos (disc_slices_curve_fit)
       - atualiza Excel (create_sheet)
       - compõe imagem final
    """

    imgProcess = img.copy()
    posEggs = findeggs(img)

    # inicializa limites para recorte final
    lMin, cMin = img.shape[0], img.shape[1]
    lMax, cMax = 0, 0

    # 2.1) montar jobs
    jobs = []
    for n, egg in enumerate(posEggs, 1):
        _, _, col, lin, larg, alt = egg
        l0 = int(lin - round(alt * 0.075))
        l1 = int(lin + round(alt * 1.075))
        c0 = int(col - round(larg * 0.075))
        c1 = int(col + round(larg * 1.075))
        rec = img[l0:l1, c0:c1].copy()
        jobs.append((n, rec.tobytes(), rec.shape, l0, l1, c0, c1, pixFactor))

    # 2.2) paralelizar SÓ processamento CPU
    maxw = max(1, multiprocessing.cpu_count() - 1)
    with ProcessPoolExecutor(max_workers=maxw) as exe:
        results = list(exe.map(cpu_process_egg, jobs))

    # 2.3) preparar diretórios de saída
    root      = Path.cwd()/'results'/Path(nomeArquivo).stem
    processed = root/'processed'
    plots     = root/'fit_plots_results'
    for d in (root, processed, plots):
        check_and_create_directory_if_not_exist(d)
    rel = open(root/'Relatorio.dad', 'a')

    # 2.4) TODO o I/O ocorre AQUI, sequencialmente
    for egg_num, l0, l1, c0, c1, out in results:
        img_proc, A, B, C, D, vol, area, x_data, y_data, pAi, pAf, pBi, pBf = out

        # --- 2.4.1) PNG e .dad -----------------------
        fn_png = processed / f'{egg_num}.png'
        cv2.imwrite(str(fn_png), img_proc)

        # vamos calcular vFormulaArea e vFormulaVolume:
        try:
            termo1 = math.acos((B/2)/D) * (D**2/math.sqrt(D**2-(B/2)**2))
            termo2 = math.acos((B/2)/C) * (C**2/math.sqrt(C**2-(B/2)**2))
            vFormulaArea = 2*math.pi*(B/2)**2 + math.pi*(B/2)*(termo1+termo2)
        except ValueError:
            vFormulaArea = 1
        try:
            vFormulaVolume = (B/2)**2 * ((2*math.pi)/3) * (D + C)
        except ValueError:
            vFormulaVolume = 1

        rel.write(f"{fn_png},{A},{B},{C},{D},{vol},{area},{vFormulaArea},{vFormulaVolume}\n")

        # --- 2.4.2) CRIA SUB‑PASTA por ovo ------------
        egg_plots = plots / str(egg_num)
        check_and_create_directory_if_not_exist(egg_plots)

        # --- 2.4.3) ajuste polinomial e gráficos -----
        x_arr = np.array(x_data)
        y_arr = np.array(y_data)
        fits, errs = disc_slices_curve_fit(
            x_arr, y_arr,
            [polynomial_curv_3, polynomial_curv_5,
             polynomial_curv_7, polynomial_curv_9, polynomial_curv_11],
            nomeArquivo, egg_num, egg_plots
        )

        # 4.4) grava imagem rotacionada, usando os pontos pAi…pBf
        rec = img[l0:l1, c0:c1]
        cSEx = int(min(pAi[0], pAf[0], pBi[0], pBf[0]))
        cSEy = int(min(pAi[1], pAf[1], pBi[1], pBf[1]))
        cIDx = int(max(pAi[0], pAf[0], pBi[0], pBf[0]))
        cIDy = int(max(pAi[1], pAf[1], pBi[1], pBf[1]))
        fn_rot = processed / f'rotate_{egg_num}.png'
        cv2.imwrite(str(fn_rot), rec[cSEx:cIDx, cSEy:cIDy])

        # 4.5) atualiza Excel
        # monta volume_results: 'Antigo' + cada ajuste
        old_slices = dict(zip(x_data, y_data))
        oldV, oldA = calcVolume(old_slices, pixFactor)
        volume_results = {'Antigo':(oldV, oldA)}
        for func, slice_d in fits.items():
            vv, aa = calcVolume(slice_d, pixFactor)
            volume_results[func] = (vv, aa)

        create_sheet(egg_plots, str(egg_num), volume_results, errs)

        # --- 2.4.6) cola no mosaico final ------------
        imgProcess[l0:l1, c0:c1] = img_proc
        lMin = min(lMin, l0)
        cMin = min(cMin, c0)
        lMax = max(lMax, l1)
        cMax = max(cMax, c1)

    rel.close()

    # recorta apenas a área com ovos
    if lMax > lMin and cMax > cMin:
        rec = imgProcess[lMin:lMax, cMin:cMax]
    else:
        rec = imgProcess  # fallback, tudo
    tL, tC, _ = adjustImageDimension(rec)
    if tL and tC:
        disp = cv2.resize(rec, (tC, tL), interpolation=cv2.INTER_LINEAR)
    else:
        disp = rec

    # 2.7) **salva o mosaico redimensionado** (e não o imgProcess gigante)
    final_fn = root / 'imagem_processada_final.png'
    cv2.imwrite(str(final_fn), disp)
    print("Processamento concluído em:", root)
    print(f"subImagens demorou: {time.perf_counter() - t0:.2f}s")

def process(frame, factor, pixFactor, egg_num, path_file_name, egg_folder_fit_plot_path):

    # Measures that will be returned
    A = B = C = D = 0

    # Uma cópia da imagem original é feita para preservar a imagem na resolução original.
    original = frame.copy()

    # Reduz a imagem para acelerar
    lin, col, ch = original.shape
    frame = resize(frame, [int(lin / factor), int(col / factor)])
    frame = np.uint8(frame * 255)

    # Segmentação HSV
    data = rgb2hsv(frame)
    channel1Min, channel1Max = 0.0, 1.0
    channel2Min, channel2Max = 0.0, 1.0
    channel3Min, channel3Max = 0.578, 1.0

    print('Creating the mask for segmentation')
    data = np.bitwise_and(
        np.bitwise_and(
            np.bitwise_and(data[:,:,0] >= channel1Min, data[:,:,0] <= channel1Max),
            np.bitwise_and(data[:,:,1] >= channel2Min, data[:,:,1] <= channel2Max)
        ),
        np.bitwise_and(data[:,:,2] >= channel3Min, data[:,:,2] <= channel3Max)
    )
    data[data > 0] = 1
    data = median(data, disk(3))
    data[data > 0] = 1

    # Encontrar bounding‐box do ovo
    print('Process to identify the bounding box that contains the egg and subsequently improve segmentation')
    linoriginal, coloriginal = data.shape
    labels = label(data)
    props = regionprops_table(labels, properties=('bbox','major_axis_length','minor_axis_length'))
    df = pd.DataFrame(props)

    fl_find_bb = False
    for _, row in df.iterrows():
        if 0.30*linoriginal <= row['major_axis_length'] <= linoriginal and \
           0.15*coloriginal <= row['minor_axis_length'] <= coloriginal:
            data = data.astype('uint8')
            data[int(row['bbox-0']):int(row['bbox-2']),
                 int(row['bbox-1']):int(row['bbox-3'])] += 1
            data[data != 2] = 0
            data[data == 2] = 255
            fl_find_bb = True
            break

    if not fl_find_bb:
        return [original, -1, -1, -1, -1]

    # Encontrar maior eixo (A) e eixo perpendicular (B)
    print('Identifying the two points that form the longest straight line')
    border_points = np.vstack(find_contours(data, 0.1))
    max_distance = 0
    pt1 = pt2 = []
    for i in range(1, len(border_points)):
        for j in range(i+1, len(border_points)):
            d = np.linalg.norm(border_points[i]*factor - border_points[j]*factor)
            if d > max_distance:
                max_distance = d
                pt1, pt2 = border_points[i]*factor, border_points[j]*factor
    if len(pt1) and len(pt2):
        original = cv2.line(original,
                            (int(pt1[1]),int(pt1[0])),
                            (int(pt2[1]),int(pt2[0])),
                            (47,141,255),2)
        A = max_distance

    # Cálculo do segundo eixo (B) e coleta de pontos para ajuste de curva
    print('Finds the angle of the line formed by the previous points and, later, finds the longest straight line')
    msRmaior = (pt1[1]*factor - pt2[1]*factor) / (pt1[0]*factor - pt2[0]*factor)
    ms = -1/msRmaior
    ms = degrees(atan(ms))
    max_distance = 0
    pt3 = pt4 = []
    dicSlicesNoFit = {}
    xdata = []
    ydata = []

    for i in range(1, len(border_points)):
        for j in range(1, len(border_points)):
            if border_points[j][0] != border_points[i][0]:
                mdRmenor = ((border_points[i][1]*factor - border_points[j][1]*factor) /
                            (border_points[i][0]*factor - border_points[j][0]*factor))
                bRmenor = border_points[i][1] - mdRmenor*border_points[i][0]
                md = degrees(atan(mdRmenor))
                if 0 <= abs(md - ms) <= 0.3:
                    d = np.linalg.norm(border_points[i]*factor - border_points[j]*factor)
                    distP1 = abs(-mdRmenor*pt1[0] + pt1[1] - bRmenor) / ((mdRmenor**2+1)**0.5)
                    original = cv2.line(original,
                                        (int(border_points[i][1]),int(border_points[i][0])),
                                        (int(border_points[j][1]),int(border_points[j][0])),
                                        (30,105,210),1)
                    dicSlicesNoFit[distP1] = d
                    xdata.append(distP1)
                    ydata.append(d)
                    if d > max_distance:
                        max_distance = d
                        pt3, pt4 = border_points[i]*factor, border_points[j]*factor

    if len(pt3) and len(pt4):
        original = cv2.line(original,
                            (int(pt3[1]),int(pt3[0])),
                            (int(pt4[1]),int(pt4[0])),
                            (0,102,0),2)
        B = max_distance

    # Linha de interseção para C e D…
    inter = lineLineIntersection(pt1,pt2,pt3,pt4)
    if inter:
        d1 = np.linalg.norm(pt1 - inter)
        d2 = np.linalg.norm(pt2 - inter)
        if d1 > d2:
            C, D = d1, d2
            original = cv2.line(original,(int(pt1[1]),int(pt1[0])),(int(inter[1]),int(inter[0])),(0,0,255),2)
            original = cv2.line(original,(int(pt2[1]),int(pt2[0])),(int(inter[1]),int(inter[0])),(255,255,255),2)
        else:
            C, D = d2, d1
            original = cv2.line(original,(int(pt2[1]),int(pt2[0])),(int(inter[1]),int(inter[0])),(0,0,255),2)
            original = cv2.line(original,(int(pt1[1]),int(pt1[0])),(int(inter[1]),int(inter[0])),(255,255,255),2)

    # =======================
    # REMOVIDO TODO ESTE BLOCO DE CÁLCULO DE CURVA E I/O:
    #
    # dic_slices_fit, curve_fit_errors = \
    #     disc_slices_curve_fit(xdata, ydata, poly_degree_functions,
    #                           path_file_name, egg_num, egg_folder_fit_plot_path)
    # volume_results = {}
    # (oldVolume, oldEggArea) = calcVolume(dicSlicesNoFit, pixFactor)
    # volume_results["Antigo"] = (oldVolume, oldEggArea)
    # for poly_function in dic_slices_fit:
    #     (poly_volume, poly_area) = calcVolume(dic_slices_fit[poly_function], pixFactor)
    #     volume_results[poly_function] = (poly_volume, poly_area)
    #     if poly_function == polynomial_curv_11:
    #         chosen_volume = poly_volume
    #         chosen_area = poly_area
    # create_sheet(egg_folder_fit_plot_path, egg_num, volume_results, curve_fit_errors)
    # =======================

    chosen_volume = 0.0
    chosen_area   = 0.0
    pAi, pAf, pBi, pBf = pt1, pt2, pt3, pt4
    # Desenha valores A, B, C, D e V
    A *= pixFactor; B *= pixFactor; C *= pixFactor; D *= pixFactor
    cv2.putText(original, f'A: {A:.3f}', (original.shape[1]-150,  30), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (47,141,255), 1)
    cv2.putText(original, f'B: {B:.3f}', (original.shape[1]-150,  50), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0,102,  0), 1)
    cv2.putText(original, f'C: {C:.3f}', (original.shape[1]-150,  70), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0,  0,255), 1)
    cv2.putText(original, f'D: {D:.3f}', (original.shape[1]-150,  90), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255,255,255), 1)
    cv2.putText(original, f'V: {chosen_volume:.3f}', (original.shape[1]-150,110), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (30,105,210),1)

    # retorna a tupla completa de dados e os vetores xdata, ydata para o ajuste ficar fora daqui
    return original, A, B, C, D, chosen_volume, chosen_area, xdata, ydata, pAi, pAf, pBi, pBf

# Function to return the intersection between two lines
def lineLineIntersection(A, B, C, D):
    # Line 01
    a1 = B[1] - A[1]
    b1 = A[0] - B[0]
    c1 = a1 * (A[0]) + b1 * (A[1])

    # Line 02
    a2 = D[1] - C[1]
    b2 = C[0] - D[0]
    c2 = a2 * (C[0]) + b2 * (C[1])

    determinant = a1 * b2 - a2 * b1

    if determinant == 0:
        return False
    else:
        x = (b2 * c1 - b1 * c2) / determinant
        y = (a1 * c2 - a2 * c1) / determinant
        return [int(x), int(y)]


def calcVolume(slices, pixFactor):
    dicSlices = sorted(slices.items())
    volume = 0
    areaLateral = 0
    # print('Inicio Ovo')
    for t in range(1, len(dicSlices)):
        elem1 = dicSlices[t]
        elem0 = dicSlices[t - 1]
        h2, rMaior = elem1
        h1, rMenor = elem0

        # print ("(",h2,",",rMaior,")")

        h = abs(h2 - h1) * pixFactor
        rMaior = (rMaior / 2) * pixFactor
        rMenor = (rMenor / 2) * pixFactor
        g = math.sqrt(h ** 2 + (rMaior - rMenor) ** 2)
        volume += ((math.pi * h) / 3) * (rMaior ** 2 + rMenor ** 2 + rMaior * rMenor)
        areaLateral += math.pi * g * (rMaior + rMenor)
    volume = volume / 100
    return [volume, areaLateral]


def disc_slices_curve_fit(x_data, y_data, polynomial_fit_degree_functions, path_file_name, egg_num, egg_folder_fit_plot_path):
    # last_function = polynomial_fit_degree_functions[-1]
    resultDiscSlices = {}
    # resultCurveError = []
    resultCurveError = {}
    for poly_fit_function in polynomial_fit_degree_functions:
        # Criar a curva de ajuste polinomial do grau escolhido na lista
        popt, pcov = curve_fit(poly_fit_function, x_data, y_data)

        # Using 'lm' method for Levenberg-Marquardt algorithm or 'trf' method for Trust Region Reflective algorithm.
        # p0 = [10, 0.1, 1, 10, 0.1, 1, 1]
        # popt, pcov = curve_fit(poly_fit_function, x_data, y_data, p0=p0, method='lm')

        y_data_fit = poly_fit_function(x_data, *popt)

        # plotando o gráfico do ajuste com os os valores de coeficiente otimizados
        plt.cla()
        plt.plot(x_data, y_data, 'b-', label='Sem ajuste (antigo)')
        plt.plot(x_data, y_data_fit, 'r-', label=f'Com ajuste: {poly_fit_function.__name__}')
        plt.suptitle(f'{Path(path_file_name).name} - Ovo {egg_num}', fontweight='bold')
        plt.title(f'Comparação: Sem ajuste X C/ ajuste ({poly_fit_function.__name__})')
        plt.xlabel("Distância X")
        plt.ylabel("Pontos de distância Y")
        plt.legend()
        # plt.show()

        # plt.savefig(rf'{__plot_results_path}/{__root_file_name}_{poly_fit_function.__name__}')
        plt.savefig(Path(egg_folder_fit_plot_path, f'plot_{poly_fit_function.__name__}'))

        # erro do ajuste polinomial
        perr = np.sqrt(np.diag(pcov))
        print(f'Error of the curve fit - {poly_fit_function.__name__}: {perr}\n')

        # Convert lists (x_data, y_data_fit) to dictionary and append it to the result list as tuple
        resultDiscSlices[poly_fit_function] = dict(zip(x_data, y_data_fit))
        # resultCurveError.append((poly_fit_function, perr))
        resultCurveError[poly_fit_function] = perr

        # if(poly_fit_function == last_function):
        # using dict() and zip() to convert lists to dictionary
        # return dict(zip(x_data, y_data_fit))
    # print(f'FINAL RESULT : {resultDiscSlices}')
    # return [np.array(resultDiscSlices), resultCurveError]
    return [resultDiscSlices, resultCurveError]

# função detecta ovos em uma imagem com base em certos critérios de área e proporção.
# Ela retorna uma lista de ovos encontrados, onde cada ovo é representado por uma lista [grupo, linha, x, y, largura, altura]
def findeggs(originalImg):
    ovos = []
    (alt, larg, ch) = originalImg.shape
    AreaTotal = alt * larg
    # preprocess the image
    gray_img = cv2.cvtColor(originalImg, cv2.COLOR_BGR2GRAY)
    # Applying 7x7 Gaussian Blur
    blurred = cv2.GaussianBlur(gray_img, (7, 7), 0)
    # Applying threshold
    threshold = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    # Apply the Component analysis function
    analysis = cv2.connectedComponentsWithStats(threshold, 4, cv2.CV_32S)
    (totalLabels, label_ids, values, centroid) = analysis
    # Loop through each component
    for i in range(1, totalLabels):
        # Area of the component
        area = values[i, cv2.CC_STAT_AREA]
        percArea = (area * 100) / AreaTotal
        aspectRatio = float(int(values[i, cv2.CC_STAT_HEIGHT]) / int(values[i, cv2.CC_STAT_WIDTH]))

        if (percArea > 0.4) and (percArea < 1) and (aspectRatio > 1.1) and (aspectRatio < 1.6):
            (col, lin) = centroid[i]
            x = values[i, cv2.CC_STAT_LEFT]
            y = values[i, cv2.CC_STAT_TOP]
            w = values[i, cv2.CC_STAT_WIDTH]
            h = values[i, cv2.CC_STAT_HEIGHT]
            elem = [0, lin, x, y, w, h]
            ovos.append(elem)
    posVet = 0
    grupo = 1
    # categoriza pela posição na linha, são classificados aquelas regiões cuja posição não linha variam abaixo de 10%
    for i in range(len(ovos)):
        if ovos[i][0] == 0:
            ovos[i][0] = grupo
            for k in range(i + 1, len(ovos)):
                if (abs(ovos[k][1] - ovos[i][1]) * 100) / ovos[i][1] < 10:
                    ovos[k][0] = grupo
            grupo += 1

    # ordena pela linha
    for i in range(0, len(ovos) - 1):
        for j in range(i + 1, len(ovos)):
            if ovos[j][0] < ovos[i][0]:
                troca = ovos[j]
                ovos[j] = ovos[i]
                ovos[i] = troca

    # ordena pela coluna
    # controle = 1
    for i in range(0, len(ovos) - 1):
        for j in range(i+1, len(ovos)):
            if ((ovos[j][0] == ovos[i][0]) and (ovos[j][2] < ovos[i][2])):
                troca = ovos[j]
                ovos[j] = ovos[i]
                ovos[i] = troca
    return ovos

"""
Representação da curva - Curva polinomial de grau 3
A função deve receber as coordenadas X e Y de seus pontos de dados como entradas e retornar os valores Y previstos para cada valor X.
"""


def polynomial_curv_3(x, a, b, c, d):
    return a * x ** 3 + b * x ** 2 + c * x + d


def polynomial_curv_5(x, a, b, c, d, e, f):
    return a * x ** 5 + b * x ** 4 + c * x ** 3 + d * x ** 2 + e * x + f


def polynomial_curv_7(x, a, b, c, d, e, f, g, h):
    return a * x ** 7 + b * x ** 6 + c * x ** 5 + d * x ** 4 + e * x ** 3 + f * x ** 2 + g * x + h


def polynomial_curv_9(x, a, b, c, d, e, f, g, h, i, j):
    return a * x ** 9 + b * x ** 8 + c * x ** 7 + d * x ** 6 + e * x ** 5 + f * x ** 4 + g * x ** 3 + h * x ** 2 + i * x + j


def polynomial_curv_11(x, a, b, c, d, e, f, g, h, i, j, k, m):
    return a * x ** 11 + b * x ** 10 + c * x ** 9 + d * x ** 8 + e * x ** 7 + f * x ** 6 + g * x ** 5 + h * x ** 4 + i * x ** 3 + j * x ** 2 + k * x + m


def piecewise_linear(x, x0, y0, k1, k2):
    return np.piecewise(x, [x <= x0], [lambda x: k1 * x + y0 - k1 * x0, lambda x: k2 * x + y0 - k2 * x0])


# def sigmoid(x, a, b, c, d):
# y = a / (1 + np.exp(-b*(x-c))) + d
# return y

# Não funcionou : Linha indo para a direita, sem seguir os pontos
def sigmoid(x, a, b, c, d):
    z = b * (x - c)
    z = np.clip(z, -500, 500)  # limit the values of the exponential term
    return a / (1 + np.exp(-z)) + d


# Resultado: Semelhante a regressão polinomial de grau 3
def gaussian(x, a, b, c, d):
    return a * np.exp(-(x - b) ** 2 / (2 * c ** 2)) + d


def get_polynomial_curv_portuguese_name(polynomial_func):
    if polynomial_func == polynomial_curv_3:
        return "Ajuste Polinomial Grau 3"
    elif polynomial_func == polynomial_curv_5:
        return "Ajuste Polinomial Grau 5"
    elif polynomial_func == polynomial_curv_7:
        return "Ajuste Polinomial Grau 7"
    elif polynomial_func == polynomial_curv_9:
        return "Ajuste Polinomial Grau 9"
    elif polynomial_func == polynomial_curv_11:
        return "Ajuste Polinomial Grau 11"
    else:
        return "Ajuste desconhecido"


# This function consists of the sum of three exponential functions with different amplitudes,
# centers, and widths, plus a constant d. You can adjust the initial parameter values to get a better fit.
# def exponential_combo(x, a1, b1, c1, a2, b2, c2, a3, b3, c3, d):
# return a1 * np.exp(-((x - b1)/c1)**2) + a2 * np.exp(-((x - b2)/c2)**2) + a3 * np.exp(-((x - b3)/c3)**2) + d
# Esta função cria ou atualiza uma planilha Excel com os resultados do volume e os erros
# de ajuste para diferentes tipos de ajuste polinomial
def create_sheet(path_file_name, egg_name, volume_results, curve_fit_errors):
    ordem_valores = ["Antigo", polynomial_curv_3, polynomial_curv_5, polynomial_curv_7, polynomial_curv_9,
                     polynomial_curv_11]

    if isinstance(path_file_name, Path):
        sheet_name_directory = Path(path_file_name.parents[2], 'Comparacao_Volume_Area.xlsx')
        file_name = path_file_name.parents[1].stem

        if os.path.exists(sheet_name_directory):
            try:
                workbook = load_workbook(sheet_name_directory)
            except Exception as e:
                print(f"Arquivo '{sheet_name_directory}' corrompido ou inválido. Criando novo workbook. Erro: {e}")
                workbook = Workbook()
        else:
            workbook = Workbook()

        sheetnames = workbook.sheetnames
        if file_name not in sheetnames:
            workbook.create_sheet(file_name)
            file_worksheet = workbook[file_name]
            file_worksheet["B1"] = "Volume - ANTIGO"
            file_worksheet["C1"] = "Area - ANTIGO"
            file_worksheet["D1"] = "Volume - POLINOM. GRAU 3"
            file_worksheet["E1"] = "Area - POLINOM. GRAU 3"
            file_worksheet["F1"] = "Erros - POLINOM. GRAU 3"
            file_worksheet["G1"] = "Volume - POLINOM. GRAU 5"
            file_worksheet["H1"] = "Area - POLINOM. GRAU 5"
            file_worksheet["I1"] = "Erros - POLINOM. GRAU 5"
            file_worksheet["J1"] = "Volume - POLINOM. GRAU 7"
            file_worksheet["K1"] = "Area - POLINOM. GRAU 7"
            file_worksheet["L1"] = "Erros - POLINOM. GRAU 7"
            file_worksheet["M1"] = "Volume - POLINOM. GRAU 9"
            file_worksheet["N1"] = "Area - POLINOM. GRAU 9"
            file_worksheet["O1"] = "Erros - POLINOM. GRAU 9"
            file_worksheet["P1"] = "Volume - POLINOM. GRAU 11"
            file_worksheet["Q1"] = "Area - POLINOM. GRAU 11"
            file_worksheet["R1"] = "Erros - POLINOM. GRAU 11"
        else:
            file_worksheet = workbook[file_name]

        # Find / Create egg row
        # is_new_egg = False
        egg_row = 0
        for row in file_worksheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value == egg_name:
                    egg_row = cell.row
        if egg_row == 0:
            max_col_row = len([cell for cell in file_worksheet["A"] if cell.value])
            egg_row = max_col_row + 2
            file_worksheet[f"A{egg_row}"] = egg_name
            # is_new_egg = True

        current_min_col = 2
        for value_type in ordem_valores:
            (volume, area) = volume_results[value_type]
            file_worksheet.cell(row=egg_row, column=current_min_col).value = volume
            file_worksheet.cell(row=egg_row, column=current_min_col + 1).value = area
            if value_type == "Antigo":
                current_min_col = current_min_col + 2
            else:
                file_worksheet.cell(row=egg_row, column=current_min_col + 2).value = ', '.join(
                    str(error) for error in curve_fit_errors[value_type])
                current_min_col = current_min_col + 3

        workbook.save(sheet_name_directory)
    else:
        print(f"input should be of type Path (pathlib) : {path_file_name}")


def exponential_combo(x, a1, b1, c1, a2, b2, c2, d):
    y1 = a1 * np.exp(-b1 * x) + c1
    y2 = a2 * np.exp(-b2 * (x - c2) ** 2) + d
    return y1 + y2

